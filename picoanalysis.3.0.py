import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import Workbook
import csv


print('Running GUI...!')
#GUI Grab file path
class Root(Tk):
    def __init__(self):
        super(Root, self).__init__()
        self.title("Browse For Your File")
        self.minsize(640, 400)

        self.labelFrame = ttk.LabelFrame(self, text = "Open the CSV to Analyze")
        self.labelFrame.grid(column = 0, row = 1, padx = 20, pady = 20)

        self.button()



    def button(self):
        self.button = ttk.Button(self.labelFrame, text = "Browse A File",command = self.fileDialog)
        self.button.grid(column = 1, row = 1)


    def fileDialog(self):
        import openpyxl
        from tkinter import ttk
        from tkinter import filedialog
        import os

        def export(coordinate,gooderlist):
            outlist=[]
            realoutlist=[]
            for data in gooderlist:
                if data[1]==coordinate:
                    outlist.append(data[2])
            for thing in outlist:
                if thing==None:
                    realoutlist.append(0)
                else: realoutlist.append(int(thing))
            return realoutlist

        def sortsec(val):
            value=int(val[1:])
            return value


        self.filename = filedialog.askopenfilename(initialdir =  "/", title = "Select A File", filetype =
        (("CSV Files","*.CSV"),("all files","*.*")) )
        self.label = ttk.Label(self.labelFrame, text = "")
        self.label.grid(column = 1, row = 2)
        self.label.configure(text = self.filename)
        #here's where you can grab the file path for use in the program.
        print('Converting CSV file to Excel Doc...')
        #Converting the CSV file to an Excel file
        wb=Workbook()
        ws=wb.active

        with open(self.filename) as csv_file:
            csv_reader=csv.reader(csv_file, delimiter=',')
            for row in csv_reader:
                ws.append(row)
        wb.save(self.filename+'.xlsx')

        print('Opening Data...')
        #Opening the data
        wb1=openpyxl.load_workbook(self.filename+'.xlsx')
        sheet1=wb1.active

        print('Formatting New Sheet...')
        #count length of sheet
        rownumber=0
        empties=0
        for row in range(1,6000):
            data=sheet1.cell(column=3, row=row).value
            rownumber+=1
            if data==None:
                empties+=1
            elif data!=None:
                empties=0
            if empties==100:
                break
        rownumber=rownumber-99

        #make the tuple and put them in a list
        tuplelist=[]
        for row in range (2,rownumber):
            datapoint1=sheet1.cell(column=1, row=row).value
            datapoint2=sheet1.cell(column=2, row=row).value
            datapoint3=sheet1.cell(column=3, row=row).value
            newtuple=(datapoint1, datapoint2, datapoint3)
            tuplelist.append(newtuple)

        #Finding number of timepoints
        totaltimes=int(tuplelist[-1][0])

        #opening new sheet
        wb2=openpyxl.Workbook()
        outsheet=wb2.active

        #making the headings
        timepoint1=[]
        for item in tuplelist:
            if item[0]=='1':
                timepoint1.append(item)

        #number of wells
        nwells=len(timepoint1)

        goodlist=[]
        for thang in range(0,totaltimes+1):
            countlist=[]
            for item in tuplelist:
                if int(item[0])==thang:
                    countlist.append(item)
            baddy=len(countlist)
            if baddy== nwells:
                for stuff in countlist:
                    goodlist.append(stuff)
            while baddy!= nwells:
                countlist.append((None, None, None))
                baddy=len(countlist)
            print('Countlist:',countlist)
            print('Number of wells recorded during timepoint,',thang,':',baddy)
        #newtotaltimes from goodlist
        totaltimes=int(goodlist[-1][0])

        #Timepoint Numbers down the left column
        timenum=1
        for row in range(3,totaltimes+3):
            outsheet.cell(column=1, row=row).value=timenum
            timenum+=1

        #grabbing the headings
        timepoint1coordinate=[]
        for guys in timepoint1:
            timepoint1coordinate.append(guys[1])

        #Ordering the headings
        timepoint1coordinate.sort()
        timepoint1coordinate.sort(key=sortsec)

        #writing the headings to the new sheet
        position=0
        for column in range (2,nwells+2):
            outsheet.cell(column=column, row=2).value=timepoint1coordinate[position]
            position+=1
        for column in range (2,nwells+2):
            position=0
            headers=outsheet.cell(column=column, row=2).value
            writelist=export(headers,goodlist)
            for row in range (3,totaltimes+3):
                outsheet.cell(column=column, row=row).value=writelist[position]
                position+=1

        outsave=self.filename+' Pico Analysis Prep.xlsx'
        wb2.save(outsave)
        print('New File Finished and Saved!')
        os.startfile(outsave)
        root.destroy()

root = Root()
root.mainloop()

#toDo: Create a an error message that tells the user where the bad data is if there isn't a full data set for a time point.

from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
import selenium.webdriver.support.expected_conditions as ec
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Font, Color, colors
from datetime import date
import time
import os
import tkinter
from tkinter import Tk
from tkinter import Label
from tkinter import LEFT
from tkinter import RIGHT
from tkinter import Entry
from tkinter import Button
import tkinter as tk

def letsgo():
    job=e1.get()
    link=e2.get()
    def valuegrab(homovalue):
        nounits=""
        for letter in homovalue:
            if letter !=" ":
                nounits+=letter
            elif letter ==" ":
                return float(nounits)
    #wait for page to load
    #blastresults=input('Paste link to BLAST results here:\n')
    #print('Opening BLAST Results Link...')
    driver= webdriver.Chrome()
    wait= WebDriverWait(driver, 1000)
    driver.get(link)
    driver.maximize_window()
    wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]')))
    #importing the primers from BLAST
    #done=input('Press "Enter" once Primer-BLASTing is complete.')
    print('Exracting Results...')
    primerlib={}
    primer1=driver.find_element_by_xpath('//*[@id="alignments"]/div[1]/table/tbody/tr[2]/td[1]').text
    primerlib[primer1]=[]
    primer2=driver.find_element_by_xpath('//*[@id="alignments"]/div[1]/table/tbody/tr[3]/td[1]').text
    primerlib[primer2]=[]
    primer3=driver.find_element_by_xpath('//*[@id="alignments"]/div[2]/table/tbody/tr[2]/td[1]').text
    primerlib[primer3]=[]
    primer4=driver.find_element_by_xpath('//*[@id="alignments"]/div[2]/table/tbody/tr[3]/td[1]').text
    primerlib[primer4]=[]
    primer5=driver.find_element_by_xpath('//*[@id="alignments"]/div[3]/table/tbody/tr[2]/td[1]').text
    primerlib[primer5]=[]
    primer6=driver.find_element_by_xpath('//*[@id="alignments"]/div[3]/table/tbody/tr[3]/td[1]').text
    primerlib[primer6]=[]
    primer7=driver.find_element_by_xpath('//*[@id="alignments"]/div[4]/table/tbody/tr[2]/td[1]').text
    primerlib[primer7]=[]
    primer8=driver.find_element_by_xpath('//*[@id="alignments"]/div[4]/table/tbody/tr[3]/td[1]').text
    primerlib[primer8]=[]
    primer9=driver.find_element_by_xpath('//*[@id="alignments"]/div[5]/table/tbody/tr[2]/td[1]').text
    primerlib[primer9]=[]
    primer10=driver.find_element_by_xpath('//*[@id="alignments"]/div[5]/table/tbody/tr[3]/td[1]').text
    primerlib[primer10]=[]
    primer11=driver.find_element_by_xpath('//*[@id="alignments"]/div[6]/table/tbody/tr[2]/td[1]').text
    primerlib[primer11]=[]
    primer12=driver.find_element_by_xpath('//*[@id="alignments"]/div[6]/table/tbody/tr[3]/td[1]').text
    primerlib[primer12]=[]
    primer13=driver.find_element_by_xpath('//*[@id="alignments"]/div[7]/table/tbody/tr[2]/td[1]').text
    primerlib[primer13]=[]
    primer14=driver.find_element_by_xpath('//*[@id="alignments"]/div[7]/table/tbody/tr[3]/td[1]').text
    primerlib[primer14]=[]
    primer15=driver.find_element_by_xpath('//*[@id="alignments"]/div[8]/table/tbody/tr[2]/td[1]').text
    primerlib[primer15]=[]
    primer16=driver.find_element_by_xpath('//*[@id="alignments"]/div[8]/table/tbody/tr[3]/td[1]').text
    primerlib[primer16]=[]
    primer17=driver.find_element_by_xpath('//*[@id="alignments"]/div[9]/table/tbody/tr[2]/td[1]').text
    primerlib[primer17]=[]
    primer18=driver.find_element_by_xpath('//*[@id="alignments"]/div[9]/table/tbody/tr[3]/td[1]').text
    primerlib[primer18]=[]
    primer19=driver.find_element_by_xpath('//*[@id="alignments"]/div[10]/table/tbody/tr[2]/td[1]').text
    primerlib[primer19]=[]
    primer20=driver.find_element_by_xpath('//*[@id="alignments"]/div[10]/table/tbody/tr[3]/td[1]').text
    primerlib[primer20]=[]

    print('Opening IDT Oligo Analyzer...')
     #Set up Web driver
    driver.get('https://www.idtdna.com/calc/analyzer/')
    driver.maximize_window()
    #wait.until(ec.presence_of_element_located((By.XPATH,'//*[@id="modal-holiday"]/div/div/div[3]/a')))
    #dumbbutton=driver.find_element_by_xpath('//*[@id="modal-holiday"]/div/div/div[3]/a')
    #dumbbutton.click()
    #wait.until(ec.visibility_of_element_located((By.ID,'UserName')))
    #sign into IDT
    username=driver.find_element_by_id('UserName')
    username.send_keys(e3.get())
    password=driver.find_element_by_id('Password')
    password.send_keys(e4.get())
    login=driver.find_element_by_id('login-button')
    login.click()
    #wait for the page to load
    wait= WebDriverWait(driver, 1000)
    wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="textarea-sequence"]')))
    #grab the Text Box
    oligoinput=driver.find_element_by_xpath('//*[@id="textarea-sequence"]')
    #Collect Hairpin, Homodimer, and Heterodimer Values for Each Primer
    print('Analyzing Primers...')
    x=0
    primerlist=[]
    for item in primerlib:
        primerlist.append(item)
    for primer in primerlib:
        #clear the textbox
        driver.find_element_by_xpath('//*[@id="textarea-sequence"]').clear()
        #enter the Primer Seq
        oligoinput.send_keys(primer)
        #Melt Temp Analyze
        Tm=driver.find_element_by_id('analyze-button')
        Tm.click()
        wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="OAResults"]/div/div[1]/div[3]/div/div/table/tbody/tr[5]/td[2]/span')))
        melted=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[1]/div[3]/div/div/table/tbody/tr[5]/td[2]/span').text
        melted=valuegrab(melted)
        primerlib[primer].append(melted)
        #Hairpin Analyze
        checkhp=driver.find_element_by_id('hairpin-button')
        checkhp.click()
        #wait for it to load the data
        wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="OAResults"]/div/div[2]/div[7]/div/div/table/tbody/tr[2]/td[3]')))
        #grab hairpin value
        hairpinvalue=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[2]/div[7]/div/div/table/tbody/tr[2]/td[3]').text
        hairpinvalue=float(hairpinvalue)
        primerlib[primer].append(hairpinvalue)
        #Homodimer Analyze
        checkhod=driver.find_element_by_xpath('//*[@id="rmenu"]/div/div[6]/button')
        checkhod.click()
        #Grab Homodimer Value
        wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="OAResults"]/div/div[3]/div[4]/span[1]')))
        homodimervalue=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[3]/div[4]/span[1]').text
        homodimervalue=valuegrab(homodimervalue)
        #Add Homodimer Value to Library
        primerlib[primer].append(homodimervalue)
        #Check out the Heterodimer Values
        if x%2 ==0:
            heterobutton=driver.find_element_by_xpath('//*[@id="rmenu"]/div/div[8]/button')
            heterobutton.click()
            wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="OAResults"]/div/div[4]/div[3]/div[4]/div/div/textarea')))
            heterotext=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[4]/div[3]/div[4]/div/div/textarea')
            heterotext.send_keys(primerlist[x+1])
            heterobutton=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[4]/div[3]/div[5]/div/div/button[2]')
            heterobutton.click()
            wait.until(ec.visibility_of_element_located((By.XPATH,'//*[@id="OAResults"]/div/div[4]/div[6]/span[1]')))
            heterovalue=driver.find_element_by_xpath('//*[@id="OAResults"]/div/div[4]/div[6]/span[1]').text
            heterovalue=valuegrab(heterovalue)
            primerlib[primer].append(heterovalue)
        x+=1
    today=date.today()
    print('Evaluating Primers...')
    #check if they're actually solid Primers
    x=0
    for primer in primerlib:
        gold=0
        if primerlib[primer][1] >=-1:
            gold+=1
        if primerlib[primer][2] >-4:
            gold+=1
        if x%2==0:
            if primerlib[primer][3] >-4:
                gold +=1
        primerlib[primer].append(gold)
        x+=1

    goldlist=[]
    for primer in primerlib:
        goldlist.append(primerlib[primer][-1])

    hitlist=[]
    x=0
    for value in range(0,20):
        if x%2==0:
            combo=goldlist[value]+goldlist[value+1]
            hitlist.append(combo)
        x+=1

    best=max(hitlist)
    bestlistpos=[]
    y=0
    for value in hitlist:
        if value == best:
            bestlistpos.append(y)
        y+=1

    print("Exporting Results...")
    #Open the Workbook
    wb=openpyxl.Workbook()
    sheet1=wb.active
    #Make the Template for the Exported Excel Sheet
    sheet1.cell(column=1, row=1).value=str(today)+" "+job+" QuickPrime Results"
    sheet1.cell(column=1, row=2).value="Sequences:"
    sheet1.cell(column=2, row=2).value="Melting Temp:"
    sheet1.cell(column=3, row=2).value="Hairpin Delta G:"
    sheet1.cell(column=4, row=2).value="Homo-dimer Delta G:"
    sheet1.cell(column=5, row=2).value="Hetero-dimer Delta G:"
    #Export the Data from the dictionary
    x=0
    for row in range(3,23):
        seqrow=sheet1.cell(column=1, row=row)
        seqrow.value=primerlist[x]
        x+=1
    x=0
    for row in range(3,23):
        temprow=sheet1.cell(column=2, row=row)
        libkey=primerlist[x]
        temprow.value=primerlib[libkey][0]
        x+=1
    x=0
    for row in range(3,23):
        hairrow=sheet1.cell(column=3, row=row)
        libkey=primerlist[x]
        hairrow.value=primerlib[libkey][1]
        x+=1
    x=0
    for row in range(3,23):
        selfrow=sheet1.cell(column=4, row=row)
        libkey=primerlist[x]
        selfrow.value=primerlib[libkey][2]
        x+=1
    x=0
    for row in range(3,23):
        if x%2==0:
            hetrow=sheet1.cell(column=5, row=row)
            libkey=primerlist[x]
            hetrow.value=primerlib[libkey][3]
        x+=1
    for pos in bestlistpos:
        pos=(pos*2)+3
        sheet1.cell(column=1, row=pos).font=Font(color=colors.BLUE)
        sheet1.cell(column=1, row=pos+1).font=Font(color=colors.BLUE)

    #Save the Workbook
    print("Saving Excel Sheet...")
    today=date.today()
    filename=str(today)+" "+job+' QuickPrime Results.xlsx'
    wb.save('D:/QuickPrimeFiles/OutFiles/'+filename)
    os.startfile('D:/QuickPrimeFiles/OutFiles'+filename)
    print('Done!')



#GUI Stuff
master = tk.Tk()
tk.Label(master, text="Enter Job Title:").grid(row=0)
tk.Label(master, text="Enter Blast Results Link:").grid(row=1)
tk.Label(master, text="IDT Username:").grid(row=2)
tk.Label(master, text="IDT Password:").grid(row=3)

e1 = tk.Entry(master)
e2 = tk.Entry(master)
e3 = tk.Entry(master)
e4 = tk.Entry(master)

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)
e3.grid(row=2, column=1)
e4.grid(row=3, column=1)

tk.Button(master, text='Quit', command=master.quit).grid(row=5, column=0, sticky=tk.W, pady=10, padx=10)
tk.Button(master, text='Begin QuickPrime', command=letsgo).grid(row=5, column=1, sticky=tk.W, pady=10, padx=10)


tk.mainloop()
